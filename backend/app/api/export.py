from flask import Blueprint, send_file, jsonify
from io import BytesIO
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..utils.db import get_db
from ..utils.decorators import jwt_required
from ..utils.operation_log import log_operation

export_bp = Blueprint("export", __name__, url_prefix="/api/export")


def create_header_style():
    return {
        "font": Font(bold=True),
        "fill": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def create_cell_style():
    return {
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    }


def set_column_widths(ws, headers, data_rows):
    """根据表头与数据设置列宽（openpyxl 列字母）。"""
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in data_rows:
            if col_idx <= len(row):
                cell_value = row[col_idx - 1]
                if cell_value is not None:
                    max_length = max(max_length, min(len(str(cell_value)), 100))
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


def safe_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    return str(value)


@export_bp.route("/excel", methods=["GET"])
@jwt_required
def export_excel():
    wb = Workbook()
    wb.remove(wb.active)

    header_style = create_header_style()
    cell_style = create_cell_style()

    db = get_db()

    try:
        ws1 = wb.create_sheet("服务器管理")
        headers1 = [
            "序号",
            "环境类型",
            "平台",
            "主机名",
            "内网IP",
            "映射IP",
            "公网IP",
            "CPU",
            "内存",
            "系统盘",
            "数据盘",
            "用途",
            "系统用户",
            "系统密码",
            "Docker用户",
            "Docker密码",
            "备注",
        ]

        with db.cursor() as cursor:
            cursor.execute("SELECT * FROM servers ORDER BY id")
            servers = cursor.fetchall()

        for col_idx, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)

        data_rows1 = []
        for idx, server in enumerate(servers, 1):
            row_data = [
                idx,
                safe_value(server.get("env_type")),
                safe_value(server.get("platform")),
                safe_value(server.get("hostname")),
                safe_value(server.get("inner_ip")),
                safe_value(server.get("mapped_ip")),
                safe_value(server.get("public_ip")),
                safe_value(server.get("cpu")),
                safe_value(server.get("memory")),
                safe_value(server.get("sys_disk")),
                safe_value(server.get("data_disk")),
                safe_value(server.get("purpose")),
                safe_value(server.get("os_user")),
                safe_value(server.get("os_password")),
                safe_value(server.get("docker_user")),
                safe_value(server.get("docker_password")),
                safe_value(server.get("remark")),
            ]
            data_rows1.append(row_data)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=idx + 1, column=col_idx, value=value)
                for key, val in cell_style.items():
                    setattr(cell, key, val)

        set_column_widths(ws1, headers1, data_rows1)

        ws2 = wb.create_sheet("服务管理")
        headers2 = [
            "序号",
            "所属服务器",
            "服务器IP",
            "环境类型",
            "分类",
            "服务名称",
            "版本",
            "内部端口",
            "映射端口",
            "外网IP",
            "内网IP",
            "备注",
        ]

        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT s.*, sv.hostname, sv.inner_ip as server_inner_ip, sv.env_type
                FROM services s
                LEFT JOIN servers sv ON s.server_id = sv.id
                ORDER BY s.id
                """
            )
            services = cursor.fetchall()

        for col_idx, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)

        data_rows2 = []
        for idx, service in enumerate(services, 1):
            row_data = [
                idx,
                safe_value(service.get("hostname")),
                safe_value(service.get("server_inner_ip")),
                safe_value(service.get("env_type")),
                safe_value(service.get("category")),
                safe_value(service.get("service_name")),
                safe_value(service.get("version")),
                safe_value(service.get("inner_port")),
                safe_value(service.get("mapped_port")),
                safe_value(service.get("public_ip")),
                safe_value(service.get("inner_ip")),
                safe_value(service.get("remark")),
            ]
            data_rows2.append(row_data)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=idx + 1, column=col_idx, value=value)
                for key, val in cell_style.items():
                    setattr(cell, key, val)

        set_column_widths(ws2, headers2, data_rows2)

        ws3 = wb.create_sheet("应用系统")
        headers3 = ["序号", "编号", "应用名称", "所属单位", "访问地址", "用户名", "密码", "备注"]

        with db.cursor() as cursor:
            cursor.execute("SELECT * FROM accounts ORDER BY id")
            apps = cursor.fetchall()

        for col_idx, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)

        data_rows3 = []
        for idx, app_row in enumerate(apps, 1):
            row_data = [
                idx,
                safe_value(app_row.get("seq_no")),
                safe_value(app_row.get("name")),
                safe_value(app_row.get("company")),
                safe_value(app_row.get("access_url")),
                safe_value(app_row.get("username")),
                safe_value(app_row.get("password")),
                safe_value(app_row.get("remark")),
            ]
            data_rows3.append(row_data)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=idx + 1, column=col_idx, value=value)
                for key, val in cell_style.items():
                    setattr(cell, key, val)

        set_column_widths(ws3, headers3, data_rows3)

        ws4 = wb.create_sheet("域名管理")
        headers4 = [
            "序号",
            "域名",
            "注册商",
            "注册日期",
            "到期日期",
            "持有者",
            "DNS服务器",
            "状态",
            "来源",
            "费用",
            "备注",
        ]

        with db.cursor() as cursor:
            cursor.execute("SELECT * FROM domains ORDER BY id")
            domains = cursor.fetchall()

        for col_idx, header in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col_idx, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)

        data_rows4 = []
        for idx, domain in enumerate(domains, 1):
            row_data = [
                idx,
                safe_value(domain.get("domain_name")),
                safe_value(domain.get("registrar")),
                safe_value(domain.get("registration_date")),
                safe_value(domain.get("expire_date")),
                safe_value(domain.get("owner")),
                safe_value(domain.get("dns_servers")),
                safe_value(domain.get("status")),
                safe_value(domain.get("source")),
                safe_value(domain.get("cost")),
                safe_value(domain.get("remark")),
            ]
            data_rows4.append(row_data)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws4.cell(row=idx + 1, column=col_idx, value=value)
                for key, val in cell_style.items():
                    setattr(cell, key, val)

        set_column_widths(ws4, headers4, data_rows4)

        ws5 = wb.create_sheet("证书管理")
        headers5 = [
            "序号",
            "域名",
            "项目",
            "类型",
            "颁发机构",
            "到期时间",
            "剩余天数",
            "品牌",
            "费用",
            "状态",
            "备注",
        ]

        with db.cursor() as cursor:
            cursor.execute("SELECT * FROM ssl_certificates ORDER BY id")
            certs = cursor.fetchall()

        for col_idx, header in enumerate(headers5, 1):
            cell = ws5.cell(row=1, column=col_idx, value=header)
            for key, value in header_style.items():
                setattr(cell, key, value)

        data_rows5 = []
        for idx, cert in enumerate(certs, 1):
            cert_type_map = {0: "自动检测", 1: "手动录入", 2: "阿里云证书"}
            cert_type = cert_type_map.get(cert.get("cert_type"), "未知")
            row_data = [
                idx,
                safe_value(cert.get("domain")),
                safe_value(cert.get("project_name")),
                cert_type,
                safe_value(cert.get("issuer")),
                safe_value(cert.get("cert_expire_time")),
                safe_value(cert.get("remaining_days")),
                safe_value(cert.get("brand")),
                safe_value(cert.get("cost")),
                safe_value(cert.get("status")),
                safe_value(cert.get("remark")),
            ]
            data_rows5.append(row_data)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws5.cell(row=idx + 1, column=col_idx, value=value)
                for key, val in cell_style.items():
                    setattr(cell, key, val)

        set_column_widths(ws5, headers5, data_rows5)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"运维数据导出_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"

        log_operation(
            module="数据导出",
            action="export",
            target_name="运维数据Excel",
            detail={"filename": filename},
        )

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        return jsonify({"code": 500, "message": f"导出失败: {str(e)}"}), 500
