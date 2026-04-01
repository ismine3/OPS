"""
从 info.xlsx 导入数据到 MySQL 数据库
"""
import openpyxl
import pymysql
from datetime import datetime
from app.config import Config
from init_db import get_connection


def import_all():
    wb = openpyxl.load_workbook('info.xlsx')
    conn = get_connection()
    cursor = conn.cursor()

    try:
        import_change_records(wb['更新记录表'], cursor)
        import_server_env(wb['数产测试环境台账'], cursor, '测试环境')
        import_server_env(wb['数产生产环境台账'], cursor, '生产环境')
        import_smart_env(wb['智慧环保生产'], cursor)
        import_water_env(wb['水电集团-安全风险智能管控'], cursor)
        import_app_systems(wb['应用系统台账表'], cursor)
        import_web_accounts(wb['web账户'], cursor)
        import_domains_certs(wb['小程序+域名'], cursor)
        conn.commit()
        print("所有数据导入完成！")
    except Exception as e:
        conn.rollback()
        print(f"导入失败: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def safe_str(val):
    """字符串清洗：处理各种空值表示"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    result = str(val).strip()
    if result in ('', '/', '-', 'NULL', 'null', 'None', '无'):
        return None
    return result if result else None


def safe_date(val):
    """日期清洗：处理各种日期格式"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        # Excel日期数字格式
        try:
            from datetime import timedelta
            base_date = datetime(1899, 12, 30)
            dt = base_date + timedelta(days=int(val))
            return dt.strftime('%Y-%m-%d')
        except:
            return None
    result = str(val).strip()
    if result in ('', '/', '-', 'NULL', 'null', 'None'):
        return None
    return result if result else None


def safe_float(val):
    """数字清洗：处理各种数字格式"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    result = str(val).strip()
    if result in ('', '/', '-', 'NULL', 'null', 'None'):
        return None
    try:
        return float(result)
    except (ValueError, TypeError):
        return None


def import_change_records(ws, cursor):
    """导入更新记录表"""
    print("导入更新记录表...")
    cursor.execute("DELETE FROM change_records")

    current_seq = None
    current_date = None
    current_modifier = None
    current_location = None

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        vals = list(row[:6])
        if all(v is None for v in vals):
            continue

        seq = vals[0]
        if seq is not None:
            current_seq = seq
            current_date = vals[1]
            current_modifier = safe_str(vals[2])
            current_location = safe_str(vals[3])

        content = safe_str(vals[4])
        remark = safe_str(vals[5])

        if content or remark:
            date_val = None
            if isinstance(current_date, datetime):
                date_val = current_date.strftime('%Y-%m-%d')

            cursor.execute(
                "INSERT INTO change_records (seq_no, change_date, modifier, location, content, remark) "
                "VALUES (%s, %s, %s, %s, %s, %s)",
                (current_seq, date_val, current_modifier, current_location, content, remark)
            )

    print(f"  更新记录导入完成")


def import_server_env(ws, cursor, env_type):
    """导入服务器环境台账（测试/生产）"""
    print(f"导入{env_type}台账...")

    current_server_id = None
    current_category = None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row[:12])
        if all(v is None for v in vals):
            continue

        platform = safe_str(vals[0])
        hostname = safe_str(vals[1])
        inner_ip = safe_str(vals[2])

        # 新服务器行
        if platform and hostname and inner_ip:
            mapped_ip = safe_str(vals[3])
            public_ip = safe_str(vals[4])

            # 根据环境类型确定密码列位置
            if env_type == '测试环境':
                docker_pwd = safe_str(vals[5])
                os_pwd = safe_str(vals[6])
            else:
                os_pwd = safe_str(vals[5])
                docker_pwd = safe_str(vals[6])

            cursor.execute(
                "INSERT INTO servers (env_type, platform, hostname, inner_ip, mapped_ip, public_ip, "
                "os_user, os_password, docker_password) "
                "VALUES (%s, %s, %s, %s, %s, %s, 'root', %s, %s)",
                (env_type, platform, hostname, inner_ip, mapped_ip, public_ip, os_pwd, docker_pwd)
            )
            current_server_id = cursor.lastrowid
            current_category = None

            # 第一行可能同时有服务信息
            svc_category = safe_str(vals[7])
            svc_name = safe_str(vals[8])
            if svc_category:
                current_category = svc_category
            if svc_name:
                version = safe_str(vals[9])
                inner_port = safe_str(vals[10])
                mapped_port = safe_str(vals[11])
                cursor.execute(
                    "INSERT INTO services (server_id, category, service_name, version, inner_port, mapped_port) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (current_server_id, current_category, svc_name, version, inner_port, mapped_port)
                )
        elif current_server_id:
            # 服务续行
            svc_category = safe_str(vals[7])
            svc_name = safe_str(vals[8])
            if svc_category and svc_category != '-':
                current_category = svc_category
            if svc_name and svc_name != '-':
                version = safe_str(vals[9])
                inner_port = safe_str(vals[10])
                mapped_port = safe_str(vals[11])
                cursor.execute(
                    "INSERT INTO services (server_id, category, service_name, version, inner_port, mapped_port) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (current_server_id, current_category, svc_name, version, inner_port, mapped_port)
                )

    print(f"  {env_type}台账导入完成")


def import_smart_env(ws, cursor):
    """导入智慧环保生产数据"""
    print("导入智慧环保生产...")

    # 这个 sheet 结构比较特殊，主要是堡垒机/服务器/VPN/管理后台/数据库的账户信息
    data_sections = []
    current_section = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row[:6])
        if all(v is None for v in vals):
            continue
        first = safe_str(vals[0])
        second = safe_str(vals[1])
        third = safe_str(vals[2])

        if first and second is None and third is None:
            current_section = first
        elif first and second and current_section:
            if current_section in ('堡垒机', '服务器', 'VPN', '管理后台', '数据库'):
                # 账户/密码行
                data_sections.append({
                    'section': current_section,
                    'account': first,
                    'password': second,
                    'ip': third
                })

    # 创建一台智慧环保服务器记录
    cursor.execute(
        "INSERT INTO servers (env_type, platform, hostname, inner_ip, os_user, os_password, remark) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s)",
        ('智慧环保', '智慧环保生产', '智慧环保生产服务器', '172.35.33.229', 'root', 'Lc13yfwpW@#!',
         '堡垒机账户/VPN账户/管理后台账户等详见数据记录')
    )
    server_id = cursor.lastrowid

    print("  智慧环保生产导入完成")


def import_water_env(ws, cursor):
    """导入水电集团-安全风险智能管控"""
    print("导入水电集团台账...")

    current_server_id = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row[:11])
        if all(v is None for v in vals):
            continue

        platform = safe_str(vals[0])
        ip = safe_str(vals[1])

        if platform and ip:
            cpu = safe_str(vals[2])
            memory = safe_str(vals[3])
            sys_disk = safe_str(vals[4])
            data_disk = safe_str(vals[5])
            purpose = safe_str(vals[6])
            account = safe_str(vals[7])
            password = safe_str(vals[8])

            cursor.execute(
                "INSERT INTO servers (env_type, platform, hostname, inner_ip, cpu, memory, "
                "sys_disk, data_disk, purpose, os_user, os_password) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                ('水电集团', platform, platform, ip, cpu, memory,
                 sys_disk, data_disk, purpose, account, password)
            )
            current_server_id = cursor.lastrowid

            svc_name = safe_str(vals[9])
            svc_port = safe_str(vals[10])
            if svc_name:
                cursor.execute(
                    "INSERT INTO services (server_id, category, service_name, inner_port) "
                    "VALUES (%s, %s, %s, %s)",
                    (current_server_id, '数据服务', svc_name, svc_port)
                )
        elif current_server_id:
            svc_name = safe_str(vals[9])
            svc_port = safe_str(vals[10])
            if svc_name:
                cursor.execute(
                    "INSERT INTO services (server_id, category, service_name, inner_port) "
                    "VALUES (%s, %s, %s, %s)",
                    (current_server_id, '数据服务', svc_name, svc_port)
                )

    print("  水电集团台账导入完成")


def import_app_systems(ws, cursor):
    """导入应用系统台账表"""
    print("导入应用系统台账表...")
    cursor.execute("DELETE FROM app_systems")

    # 表头在第2行，数据从第3行开始
    # 列顺序: 序号(0)|名称(1)|单位(2)|架构(3)|登录链接(4)|账号(5)|密码(6)|备注(7)|更新时间(8)
    # 映射: seq_no=vals[0], name=vals[1], company=vals[2], access_url=vals[4](跳过架构), username=vals[5], password=vals[6], remark=vals[7]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row[:9])  # 读取前9列
        if all(v is None for v in vals):
            continue

        seq_no = safe_str(vals[0])
        name = safe_str(vals[1])
        if not name:
            continue

        company = safe_str(vals[2])
        # 跳过架构列(vals[3])，直接使用登录链接
        access_url = safe_str(vals[4])
        username = safe_str(vals[5])
        password = safe_str(vals[6])
        remark = safe_str(vals[7])
        # 跳过更新时间(vals[8])

        cursor.execute(
            "INSERT INTO app_systems (seq_no, name, company, access_url, username, password, remark) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (seq_no, name, company, access_url, username, password, remark)
        )

    print("  应用系统台账导入完成")


def import_web_accounts(ws, cursor):
    """导入web账户Sheet到app_systems表（追加数据）"""
    print("导入web账户...")

    # 获取当前最大序号，用于自动生成seq_no
    cursor.execute("SELECT MAX(CAST(seq_no AS UNSIGNED)) FROM app_systems WHERE seq_no IS NOT NULL")
    result = cursor.fetchone()
    max_seq = result[0] if result and result[0] else 0
    current_seq = max_seq

    current_company = None

    # 数据结构: 分组式，分组名在某些行的第1列
    # 数据列: 应用名称(0) | 访问链接(1) | 账户(2) | 密码(3)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        vals = list(row[:4])

        # 检查是否为空行
        if all(v is None for v in vals):
            continue

        # 判断是否为分组行：只有第1列有值，其余列为空
        first_col = safe_str(vals[0])
        second_col = vals[1] if len(vals) > 1 else None
        third_col = vals[2] if len(vals) > 2 else None
        fourth_col = vals[3] if len(vals) > 3 else None

        is_group_row = (first_col and
                       (second_col is None or safe_str(second_col) is None) and
                       (third_col is None or safe_str(third_col) is None) and
                       (fourth_col is None or safe_str(fourth_col) is None))

        if is_group_row:
            # 这是分组名（公司名）
            current_company = first_col
            continue

        # 数据行
        name = safe_str(vals[0])
        if not name:
            continue

        access_url = safe_str(vals[1])
        username = safe_str(vals[2])
        password = safe_str(vals[3])

        # 自动生成序号
        current_seq += 1

        cursor.execute(
            "INSERT INTO app_systems (seq_no, name, company, access_url, username, password, remark) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (current_seq, name, current_company, access_url, username, password, None)
        )

    print(f"  web账户导入完成，共导入 {current_seq - max_seq} 条记录")


def import_domains_certs(ws, cursor):
    """导入域名与证书表（小程序+域名Sheet）"""
    print("导入域名与证书表...")
    cursor.execute("DELETE FROM domains_certs")

    # 表头在第7行，数据从第8行开始
    # 列顺序: 序号(0)|项目(1)|项目子类(2)|主体(3)|购买时间(4)|到期时间(5)|费用(6)|截止有效期(7)
    current_category = None

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        vals = list(row[:8])
        if all(v is None for v in vals):
            continue

        seq_no = safe_str(vals[0])
        project_main = safe_str(vals[1])  # 项目
        project_sub = safe_str(vals[2])   # 项目子类
        entity = safe_str(vals[3])

        # 根据项目名判断分类
        if project_main:
            if '公众' in project_main or '小程序' in project_main:
                current_category = '公众平台'
            elif '域名' in project_main:
                current_category = '域名'
            elif 'SSL' in project_main or '证书' in project_main:
                current_category = 'SSL证书'
            else:
                current_category = project_main

        # 取项目名（优先项目子类，如果没有则使用项目）
        project = project_sub if project_sub else project_main
        if not project:
            continue

        purchase_date = safe_date(vals[4])
        expire_date = safe_date(vals[5])
        cost = safe_float(vals[6])
        remaining_days = safe_str(vals[7])

        cursor.execute(
            "INSERT INTO domains_certs (seq_no, category, project, entity, purchase_date, "
            "expire_date, cost, remaining_days) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            (seq_no, current_category, project, entity, purchase_date,
             expire_date, cost, remaining_days)
        )

    print("  域名与证书导入完成")


if __name__ == '__main__':
    import_all()
